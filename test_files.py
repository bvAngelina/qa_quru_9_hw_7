import os.path
import pytest
import zipfile
from pypdf import PdfReader
from io import TextIOWrapper
from openpyxl import load_workbook
import csv

CURRENT_FILE = os.path.abspath(__file__)
CURRENT_DIR = os.path.dirname(CURRENT_FILE)
path = os.path.join(CURRENT_DIR, "tmp")
resources = os.path.join(CURRENT_DIR, "resources")
zip_path = os.path.join(resources, "test_archive.zip")


@pytest.fixture
def archive_files():
    if not os.path.exists(resources):
        os.mkdir(resources)
    file_dir = os.listdir(path)
    with zipfile.ZipFile(zip_path, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        for file in file_dir:
            add_file = os.path.join(path, file)
            zf.write(add_file, file)
    yield
    os.remove(zip_path)


def test_check_files(archive_files):
    with zipfile.ZipFile("resources/test_archive.zip") as zip_file:
        print(zip_file.namelist())
        with zip_file.open("file1.pdf") as pdf_file:
            reader = PdfReader(pdf_file)
            text = reader.pages[0].extract_text()
            assert "День занялся тусклый, серый." in text

        with zip_file.open("file2.csv", "r") as csv_file:
            reader = list(csv.reader(TextIOWrapper(csv_file)))
            print(reader[1][0])
            assert "123;текст;text" == reader[1][0]

        with zip_file.open("file3.xlsx") as xlsx:
            workbook = load_workbook(xlsx)
            sheet = workbook.active
            print(sheet.cell(row=1, column=1))
            assert sheet.cell(row=1, column=1).value == 123
